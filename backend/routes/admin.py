from flask import Blueprint, request, jsonify, send_file
from models.student import Student
from models.event import Event
from models.attendance import Attendance
from models.unauthorized_log import UnauthorizedLog
from utils.qr_generator import QRGenerator
from utils.sanitizer import InputSanitizer
import jwt
from config import Config
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime

admin_bp = Blueprint('admin', __name__)

def verify_admin_token(token):
    """Verify admin JWT token"""
    try:
        payload = jwt.decode(token, Config.JWT_SECRET_KEY, algorithms=['HS256'])
        if payload.get('type') != 'admin':
            return None, 'Invalid token type'
        return payload, None
    except jwt.ExpiredSignatureError:
        return None, 'Token has expired'
    except jwt.InvalidTokenError:
        return None, 'Invalid token'


@admin_bp.route('/dashboard', methods=['GET'])
def get_dashboard():
    """Get admin dashboard statistics"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        # Fetch all data in a few bulk queries (avoids N+1 per event)
        from supabase_client import get_supabase_client, reset_supabase_client
        
        def _fetch_dashboard_data():
            sb = get_supabase_client()
            students_res = sb.table('students').select('id, department, registered_events').execute()
            all_students = students_res.data or []

            events_res = sb.table('events').select('id, event_name').execute()
            all_events_raw = events_res.data or []

            att_res = sb.table('attendance').select('event_id').execute()
            all_attendance = att_res.data or []
            return all_students, all_events_raw, all_attendance

        try:
            all_students, all_events_raw, all_attendance = _fetch_dashboard_data()
        except Exception as conn_err:
            # Retry once with a fresh client on connection errors
            reset_supabase_client()
            all_students, all_events_raw, all_attendance = _fetch_dashboard_data()

        total_students = len(all_students)
        total_events = len(all_events_raw)

        # Build attendance count per event
        att_count_map = {}
        for a in all_attendance:
            eid = str(a.get('event_id', ''))
            att_count_map[eid] = att_count_map.get(eid, 0) + 1

        # Build registration count per event from students
        reg_count_map = {}
        for student in all_students:
            for eid in (student.get('registered_events') or []):
                key = str(eid)
                reg_count_map[key] = reg_count_map.get(key, 0) + 1

        # Build event registrations list
        event_registrations = []
        for event in all_events_raw:
            event_id = str(event.get('id', ''))
            total_registered = reg_count_map.get(event_id, 0)
            attendance_count = att_count_map.get(event_id, 0)
            event_registrations.append({
                'event_id': event_id,
                'event_name': event['event_name'],
                'total_registered': total_registered,
                'total_attended': attendance_count,
                'attendance_percentage': round((attendance_count / total_registered * 100), 2) if total_registered > 0 else 0
            })

        # Department-wise statistics from in-memory data
        dept_map = {}
        for student in all_students:
            dept = student.get('department') or 'Unknown'
            dept_map[dept] = dept_map.get(dept, 0) + 1
        department_stats = sorted(
            [{'department': k, 'count': v} for k, v in dept_map.items()],
            key=lambda x: -x['count']
        )

        # Get unauthorized scans count
        unauthorized_count = UnauthorizedLog.count_unauthorized()
        
        return jsonify({
            'success': True,
            'dashboard': {
                'total_students': total_students,
                'total_events': total_events,
                'event_registrations': event_registrations,
                'department_stats': department_stats,
                'unauthorized_scans': unauthorized_count
            }
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to fetch dashboard: {str(e)}'}), 500


@admin_bp.route('/generate-qr/<event_id>', methods=['POST'])
def generate_qr(event_id):
    """Generate QR code for an event"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        # Verify event exists
        event = Event.find_by_id(event_id)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found'}), 404
        
        # Generate QR code with event name
        qr_result = QRGenerator.generate_qr_code(event_id, event['event_name'])
        
        return jsonify({
            'success': True,
            'qr_code': qr_result['qr_image'],
            'event_name': event['event_name'],
            'secure_token': qr_result['secure_token']
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'QR generation failed: {str(e)}'}), 500


@admin_bp.route('/attendance-report/<event_id>', methods=['GET'])
def get_attendance_report(event_id):
    """Get attendance report for an event"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        # Verify event exists
        event = Event.find_by_id(event_id)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found'}), 404
        
        # Get all students registered for this event
        registered_students = Student.get_students_by_event(event_id)
        
        # Get attendance records
        attendance_records = Attendance.get_event_attendance(event_id)
        
        # Create attendance map
        attendance_map = {}
        for record in attendance_records:
            attendance_map[str(record['student_id'])] = record['marked_at']
        
        # Build complete student list with attendance status
        student_list = []
        for student in registered_students:
            student_id = InputSanitizer.get_id(student)
            is_present = student_id in attendance_map
            
            student_list.append({
                'student_name': student['name'],
                'register_number': student['register_number'],
                'department': student['department'],
                'course': student.get('course', 'N/A'),
                'year': student['year'],
                'status': 'Present' if is_present else 'Absent',
                'marked_at': attendance_map[student_id].isoformat() if is_present and hasattr(attendance_map[student_id], 'isoformat') else (attendance_map[student_id] if is_present else None)
            })
        
        # Sort by register number
        student_list.sort(key=lambda x: x['register_number'])
        
        total_registered = len(registered_students)
        total_attended = len(attendance_map)
        total_absent = total_registered - total_attended
        
        return jsonify({
            'success': True,
            'event_name': event['event_name'],
            'total_registered': total_registered,
            'total_attended': total_attended,
            'total_absent': total_absent,
            'attendance_percentage': round((total_attended / total_registered * 100), 2) if total_registered > 0 else 0,
            'student_list': student_list
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to fetch report: {str(e)}'}), 500


@admin_bp.route('/export-attendance/<event_id>', methods=['GET'])
def export_attendance(event_id):
    """Export attendance report to Excel with all registered students"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        # Verify event exists
        event = Event.find_by_id(event_id)
        if not event:
            return jsonify({'success': False, 'message': 'Event not found'}), 404
        
        # Get all students registered for this event
        registered_students = Student.get_students_by_event(event_id)
        
        # Get attendance records
        attendance_records = Attendance.get_event_attendance(event_id)
        
        # Create attendance map
        attendance_map = {}
        for record in attendance_records:
            attendance_map[str(record['student_id'])] = record['marked_at']
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add title and event info
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Attendance Report - {event['event_name']}"
        ws['A1'].font = ws['A1'].font.copy(size=16, bold=True)
        
        # Add headers in row 3
        headers = ['S.No', 'Register Number', 'Name', 'Department', 'Course', 'Year', 'Status', 'Attendance Time']
        ws.append([])  # Empty row
        ws.append(headers)
        
        # Style headers
        for cell in ws[3]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(fgColor="4472C4", patternType="solid")
            cell.font = cell.font.copy(color="FFFFFF")
        
        # Sort students by register number
        registered_students.sort(key=lambda x: x.get('register_number', ''))
        
        # Add data
        for idx, student in enumerate(registered_students, 1):
            student_id = InputSanitizer.get_id(student)
            is_present = student_id in attendance_map
            
            # Format attendance time - handle both datetime objects and ISO strings
            marked_at = attendance_map[student_id] if is_present else None
            if marked_at:
                if hasattr(marked_at, 'strftime'):
                    # MongoDB datetime object
                    attendance_time = marked_at.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    # Supabase ISO string - convert to readable format
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(marked_at.replace('Z', '+00:00'))
                        attendance_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        attendance_time = str(marked_at)
            else:
                attendance_time = '-'
            
            row = [
                idx,
                student['register_number'],
                student['name'],
                student['department'],
                student.get('course', 'N/A'),
                student['year'],
                'Present' if is_present else 'Absent',
                attendance_time
            ]
            ws.append(row)
            
            # Color code rows
            row_num = idx + 3
            if is_present:
                for cell in ws[row_num]:
                    cell.fill = cell.fill.copy(fgColor="C6EFCE", patternType="solid")
            else:
                for cell in ws[row_num]:
                    cell.fill = cell.fill.copy(fgColor="FFC7CE", patternType="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Send file
        filename = f"attendance_{event['event_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500


@admin_bp.route('/student/<student_id>', methods=['PUT'])
def update_student(student_id):
    """Update student information (admin only)"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        if error:
            return jsonify({'success': False, 'message': error}), 401

        data = request.get_json() or {}
        allowed_fields = {'name': InputSanitizer.sanitize_name,
                          'department': InputSanitizer.sanitize_string,
                          'course': InputSanitizer.sanitize_string,
                          'year': InputSanitizer.sanitize_string,
                          'email': InputSanitizer.sanitize_email,
                          'phone_number': InputSanitizer.sanitize_phone}

        update_payload = {}
        for field, sanitizer in allowed_fields.items():
            if field in data and data[field] is not None:
                update_payload[field] = sanitizer(str(data[field]))

        if not update_payload:
            return jsonify({'success': False, 'message': 'No valid fields provided'}), 400

        update_payload['updated_at'] = datetime.utcnow().isoformat()

        from supabase_client import get_supabase_client
        sb = get_supabase_client()
        sb.table('students').update(update_payload).eq('id', student_id).execute()

        return jsonify({'success': True, 'message': 'Student updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Update failed: {str(e)}'}), 500


@admin_bp.route('/student/<student_id>', methods=['DELETE'])
def delete_student(student_id):
    """Delete a student and their attendance records (admin only)"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        if error:
            return jsonify({'success': False, 'message': error}), 401

        from supabase_client import get_supabase_client
        sb = get_supabase_client()
        # Remove attendance records first to avoid orphans
        sb.table('attendance').delete().eq('student_id', student_id).execute()
        sb.table('students').delete().eq('id', student_id).execute()

        return jsonify({'success': True, 'message': 'Student deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Delete failed: {str(e)}'}), 500


@admin_bp.route('/export-all-attendance', methods=['GET'])
def export_all_attendance():
    """Export attendance summary for all events to a multi-sheet Excel file"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        if error:
            return jsonify({'success': False, 'message': error}), 401

        from supabase_client import get_supabase_client
        sb = get_supabase_client()
        events_res = sb.table('events').select('id, event_name, date').execute()
        all_events = events_res.data or []

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Summary'

        summary_headers = ['S.No', 'Event Name', 'Date', 'Registered', 'Attended', 'Absent', 'Attendance %']
        ws_summary.append(summary_headers)
        for cell in ws_summary[1]:
            cell.font = cell.font.copy(bold=True, color='FFFFFF')
            cell.fill = cell.fill.copy(fgColor='4472C4', patternType='solid')

        for idx, event in enumerate(all_events, 1):
            event_id = str(event['id'])
            registered_students = Student.get_students_by_event(event_id)
            attendance_records = Attendance.get_event_attendance(event_id)
            attendance_map = {str(r['student_id']): r['marked_at'] for r in attendance_records}

            total_reg = len(registered_students)
            total_att = len(attendance_map)
            total_abs = total_reg - total_att
            pct = round(total_att / total_reg * 100, 1) if total_reg > 0 else 0

            ws_summary.append([idx, event['event_name'], event.get('date', 'N/A'),
                               total_reg, total_att, total_abs, f'{pct}%'])

            # Per-event detail sheet (Excel sheet names max 31 chars)
            sheet_name = event['event_name'][:31]
            ws = wb.create_sheet(title=sheet_name)
            detail_headers = ['S.No', 'Register Number', 'Name', 'Department', 'Course', 'Year', 'Status', 'Attendance Time']
            ws.append(detail_headers)
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True, color='FFFFFF')
                cell.fill = cell.fill.copy(fgColor='4472C4', patternType='solid')

            registered_students.sort(key=lambda x: x.get('register_number', ''))
            for i, student in enumerate(registered_students, 1):
                student_id = InputSanitizer.get_id(student)
                is_present = student_id in attendance_map
                marked_at = attendance_map.get(student_id)
                if marked_at:
                    try:
                        dt = datetime.fromisoformat(marked_at.replace('Z', '+00:00'))
                        attendance_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        attendance_time = str(marked_at)
                else:
                    attendance_time = '-'

                ws.append([i, student['register_number'], student['name'], student['department'],
                           student.get('course', 'N/A'), student['year'],
                           'Present' if is_present else 'Absent', attendance_time])
                color = 'C6EFCE' if is_present else 'FFC7CE'
                for cell in ws[i + 1]:
                    cell.fill = cell.fill.copy(fgColor=color, patternType='solid')

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        filename = f"all_events_attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500


@admin_bp.route('/unauthorized-logs', methods=['GET'])
def get_unauthorized_logs():
    """Get all unauthorized scan attempts"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        logs = UnauthorizedLog.get_all_logs()
        
        logs_list = []
        for log in logs:
            logs_list.append({
                'id': InputSanitizer.get_id(log),
                'ip_address': log.get('ip_address', 'Unknown'),
                'user_agent': log.get('user_agent', 'Unknown'),
                'scan_data': log.get('scan_data'),
                'scanned_at': log['scanned_at'].isoformat() if hasattr(log['scanned_at'], 'isoformat') else log['scanned_at']
            })
        
        return jsonify({
            'success': True,
            'total_logs': len(logs_list),
            'logs': logs_list
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to fetch logs: {str(e)}'}), 500


@admin_bp.route('/attendance-statistics', methods=['GET'])
def get_attendance_statistics():
    """Get comprehensive attendance statistics"""
    try:
        # Verify admin token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        payload, error = verify_admin_token(token)
        
        if error:
            return jsonify({'success': False, 'message': error}), 401
        
        # Get attendance statistics
        stats = Attendance.get_attendance_statistics()
        
        statistics = []
        for stat in stats:
            statistics.append({
                'event_name': stat.get('event_name', 'Unknown'),
                'department': stat.get('department', 'Unknown'),
                'count': stat['count']
            })
        
        return jsonify({
            'success': True,
            'statistics': statistics
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to fetch statistics: {str(e)}'}), 500
